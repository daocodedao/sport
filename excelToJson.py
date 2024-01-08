# -*- coding: utf-8 -*-
import os
import sys

import openpyxl
import json
import io

# 将json保存为文件
def save2json(jd, json_file_name):
    file = io.open(json_file_name, 'w', encoding='utf-8')
    # 把对象转化为json对象
    # indent: 参数根据数据格式缩进显示，读起来更加清晰
    # ensure_ascii = True：默认输出ASCII码，如果把这个该成False, 就可以输出中文。
    txt = json.dumps(jd, indent=2, ensure_ascii=False)
    file.write(txt)
    file.close()

# excel表格转json文件
def excel2json(excel_file, json_file_name):
    # 加载工作薄
    book = openpyxl.load_workbook(excel_file)
    # 获取sheet页
    sheet = book["Sheet1"]
    # 行数
    max_row = sheet.max_row
    # 列数
    max_column = sheet.max_column
    # print("max_row: %d, max_column: %d" % (max_row, max_column))
    # 结果，数组存储
    result = []
    heads = []
    # 解析表头
    for column in range(max_column):
        # 读取的话行列是从（1，1）开始
        heads.append(sheet.cell(1, column + 1).value)
    # 遍历每一行
    for row in range(max_row):
        if row == 0:
            continue
        one_line = {}
        for column in range(max_column):
            # 读取第二行开始每一个数据
            k = heads[column]
            cell = sheet.cell(row + 1, column + 1)
            value = cell.value
            one_line[k] = value
        print(one_line)
        result.append(one_line)
    book.close()
    # 将json保存为文件
    save2json(result, json_file_name)


# #main
# if '__main__' == __name__:
#      excel2json(u'data.xlsx', 'res.json')

